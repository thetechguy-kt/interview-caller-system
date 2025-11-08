from tkinter import messagebox, font as tkfont
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import tkinter as tk
import os
from datetime import datetime
import qrcode
import shutil

EXCEL_FILE = "candidate_list.xlsx"
TICKET_FOLDER = "Tickets"
CONFIG_FOLDER = "config"
os.makedirs(CONFIG_FOLDER, exist_ok=True)
DATE_TRACK_FILE = os.path.join(CONFIG_FOLDER, "last_ticket_date.txt")

# --- Utility function to pick best font ---
def pick_preferred_font(root):
    preferred_fonts = ["Montserrat", "Aptos", "Segoe UI", "Helvetica", "Arial"]
    available = list(tkfont.families(root))
    for f in preferred_fonts:
        if f in available:
            return f
    return "TkDefaultFont"

# --- Register Montserrat or Aptos if available ---
def register_pdf_font():
    possible_fonts = [
        ("Montserrat", "C:\\Windows\\Fonts\\Montserrat-Regular.ttf"),
        ("Aptos", "C:\\Windows\\Fonts\\Aptos.ttf"),
    ]
    for name, path in possible_fonts:
        if os.path.exists(path):
            pdfmetrics.registerFont(TTFont(name, path))
            return name
    # fallback
    return "Helvetica"

class InterviewCandidatePOS:
    def __init__(self, root):
        self.root = root
        self.root.title("KTech Candidate POS")
        self.set_window_size(420, 480)

        self.bg_color = "#121217"
        self.fg_color = "#E0E6F1"
        self.accent_color = "#00FFFF"
        self.button_bg = "#1F1F2B"
        self.button_hover_bg = "#00CED1"
        self.entry_bg = "#1E1E2F"
        self.entry_fg = "#E0E6F1"
        self.root.configure(bg=self.bg_color)

        self.font_family = pick_preferred_font(root)
        self.pdf_font = register_pdf_font()

        self.today = datetime.now().strftime("%Y-%m-%d")
        self.check_and_reset_daily()

        self.main_frame = tk.Frame(root, bg=self.bg_color)
        self.main_frame.pack(fill='both', expand=True)
        self.main_frame.rowconfigure(0, weight=1)
        self.main_frame.rowconfigure(1, weight=0)
        self.main_frame.columnconfigure(0, weight=1)

        self.input_frame = tk.Frame(self.main_frame, bg=self.bg_color, padx=20, pady=20)
        self.input_frame.grid(row=0, column=0, sticky='nsew')

        self.button_frame = tk.Frame(self.main_frame, bg=self.bg_color, padx=20, pady=10)
        self.button_frame.grid(row=1, column=0, sticky='ew')

        for i in range(6):
            self.input_frame.rowconfigure(i, weight=0)
        self.input_frame.columnconfigure(0, weight=0)
        self.input_frame.columnconfigure(1, weight=1)

        self.title_label = tk.Label(
            self.input_frame,
            text="🎓 KTech Interview Candidate POS",
            font=(self.font_family, 18, "bold"),
            bg=self.bg_color, fg=self.accent_color
        )
        self.title_label.grid(row=0, column=0, columnspan=2, pady=(0, 25))

        self.name_label = tk.Label(
            self.input_frame, text="Candidate Name:",
            bg=self.bg_color, fg=self.fg_color, font=(self.font_family, 12)
        )
        self.name_label.grid(row=1, column=0, sticky='w', pady=5)
        self.name_entry = tk.Entry(
            self.input_frame, bg=self.entry_bg, fg=self.entry_fg,
            insertbackground=self.accent_color, font=(self.font_family, 12),
            relief="flat", highlightthickness=2,
            highlightbackground="#2F2F3F", highlightcolor=self.accent_color
        )
        self.name_entry.grid(row=1, column=1, sticky='ew', pady=5)

        self.contact_label = tk.Label(
            self.input_frame, text="Contact Number:",
            bg=self.bg_color, fg=self.fg_color, font=(self.font_family, 12)
        )
        self.contact_label.grid(row=2, column=0, sticky='w', pady=5)
        self.contact_number_entry = tk.Entry(
            self.input_frame, bg=self.entry_bg, fg=self.entry_fg,
            insertbackground=self.accent_color, font=(self.font_family, 12),
            relief="flat", highlightthickness=2,
            highlightbackground="#2F2F3F", highlightcolor=self.accent_color
        )
        self.contact_number_entry.grid(row=2, column=1, sticky='ew', pady=5)

        self.ticket_label = tk.Label(
            self.input_frame, text=f"Entry No: {getattr(self, 'ticket_number', 0)}",
            font=(self.font_family, 22, "bold"), fg=self.accent_color, bg=self.bg_color
        )
        self.ticket_label.grid(row=4, column=0, columnspan=2, pady=20)

        self.btn_generate = tk.Button(
            self.button_frame, text="Generate Entry Pass",
            font=(self.font_family, 14, "bold"),
            bg=self.button_bg, fg=self.accent_color,
            activebackground=self.button_hover_bg, activeforeground="#121217",
            relief="flat", command=self.generate_ticket, cursor="hand2"
        )
        self.btn_generate.pack(fill='x', pady=8)
        self.add_hover_effect(self.btn_generate, self.button_bg, self.button_hover_bg, self.accent_color, "#121217")

        self.btn_reset = tk.Button(
            self.button_frame, text="Reset Counter",
            font=(self.font_family, 12, "bold"),
            bg="#8B0000", fg="white",
            activebackground="#B22222", activeforeground="#f0f0f0",
            relief="flat", command=self.reset_counter, cursor="hand2"
        )
        self.btn_reset.pack(fill='x')
        self.add_hover_effect(self.btn_reset, "#8B0000", "#B22222", "white", "#f0f0f0")

        self.setup_excel()

    def add_hover_effect(self, widget, bg_normal, bg_hover, fg_normal, fg_hover):
        def on_enter(e):
            widget['background'] = bg_hover
            widget['foreground'] = fg_hover
            widget['font'] = (self.font_family, widget['font'][1], "bold")
        def on_leave(e):
            widget['background'] = bg_normal
            widget['foreground'] = fg_normal
            widget['font'] = (self.font_family, widget['font'][1])
        widget.bind("<Enter>", on_enter)
        widget.bind("<Leave>", on_leave)

    def set_window_size(self, width, height):
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")
        self.root.minsize(400, 450)

    def check_and_reset_daily(self):
        if not os.path.exists(DATE_TRACK_FILE):
            with open(DATE_TRACK_FILE, 'w') as f:
                f.write(self.today)
            self.ticket_number = 0
        else:
            with open(DATE_TRACK_FILE, 'r') as f:
                last_date = f.read().strip()

            if last_date != self.today:
                self.ticket_number = 0
                with open(DATE_TRACK_FILE, 'w') as f:
                    f.write(self.today)
                self.reset_excel_file()
            else:
                self.ticket_number = self.get_last_ticket_number()

    def reset_excel_file(self):
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            ws.delete_rows(2, ws.max_row)
            wb.save(EXCEL_FILE)

    def get_last_ticket_number(self):
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == self.today:
                    count += 1
            return count
        return 0

    def setup_excel(self):
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            headers = ["Date", "Day", "Time", "Candidate Name", "Contact Number", "Entry No"]
            ws.append(headers)
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="003F5C", end_color="003F5C", fill_type="solid")
                ws.column_dimensions[chr(64 + col)].width = max(len(header) + 5, 15)
            wb.save(EXCEL_FILE)

    def generate_ticket(self):
        name = self.name_entry.get().strip()
        contact_number = self.contact_number_entry.get().strip()
        if not name:
            messagebox.showwarning("Input Required", "Please enter the candidate's name.")
            return
        if not contact_number:
            messagebox.showwarning("Input Required", "Please enter the contact number.")
            return

        now = datetime.now()
        date = now.strftime("%Y-%m-%d")
        day = now.strftime("%A")
        time = now.strftime("%H:%M:%S")
        file_time = now.strftime("%H-%M-%S")

        self.ticket_number += 1
        self.ticket_label.config(text=f"Entry No: {self.ticket_number}")

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([date, day, time, name, contact_number, self.ticket_number])
        new_row = ws.max_row
        for col in range(1, 7):
            ws.cell(row=new_row, column=col).alignment = Alignment(horizontal='center')
        wb.save(EXCEL_FILE)

        folder_name = os.path.join(TICKET_FOLDER, f"{date} - Entries")
        os.makedirs(folder_name, exist_ok=True)
        daily_excel_path = os.path.join(folder_name, f"candidate_list_{date}.xlsx")
        shutil.copy(EXCEL_FILE, daily_excel_path)

        safe_name = name.replace(" ", "_")
        pdf_filename = f"Entry_{self.ticket_number}_{safe_name}_{file_time}.pdf"
        pdf_path = os.path.join(folder_name, pdf_filename)

        self.create_ticket_pdf(pdf_path, name, contact_number, self.ticket_number, date, day, time)

        self.name_entry.delete(0, tk.END)
        self.contact_number_entry.delete(0, tk.END)

        messagebox.showinfo("Success", f"Entry No {self.ticket_number} generated for {name}.")
        if messagebox.askyesno("Print Entry Pass", "Do you want to print the pass now?"):
            try:
                os.startfile(pdf_path, "print")
            except Exception as e:
                messagebox.showerror("Printing Error", f"Could not print ticket: {e}")

    def create_ticket_pdf(self, filepath, name, contact_number, entry_no, date, day, time):
        width = 8 * cm
        height = 8 * cm
        c = canvas.Canvas(filepath, pagesize=(width, height))
        qr_text = f"Entry No: {entry_no}\nName: {name}\nContact: {contact_number}\nDate: {date} ({day})\nTime: {time}"
        qr_img = qrcode.make(qr_text)
        qr_temp = "temp_qr.png"
        qr_img.save(qr_temp)

        c.setFont(self.pdf_font, 16)
        c.setFillColorRGB(0, 0, 0)
        c.drawCentredString(width / 2, height - 30, "KTech")

        c.setFont(self.pdf_font, 10)
        c.drawCentredString(width / 2, height - 50, f"{date} ({day}) | {time}")

        c.setFont(self.pdf_font, 10)
        c.drawString(20, height - 80, f"Name: {name}")
        c.drawString(20, height - 110, f"Number: {contact_number}")
        c.drawString(20, height - 140, f"Entry No: {entry_no}")

        c.drawImage(qr_temp, width - 90, 20, width=70, height=70)

        c.setFont(self.pdf_font, 8)
        c.drawCentredString(width / 2, 10, "Scan for interview info")

        c.showPage()
        c.save()
        os.remove(qr_temp)

    def reset_counter(self):
        if messagebox.askyesno("Confirm Reset", "Are you sure you want to reset the entry number?"):
            self.ticket_number = 0
            self.ticket_label.config(text="Entry No: 0")
            self.reset_excel_file()

if __name__ == "__main__":
    root = tk.Tk()
    root.configure(bg="#121217")
    app = InterviewCandidatePOS(root)
    root.mainloop()
