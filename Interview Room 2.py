import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
import json
import os
from datetime import datetime
import tkinter.font as tkfont

# Constants
EXCEL_FILE = "candidate_list.xlsx"
STATE_FILE = "queue_state.json"
COUNTER_NAME = "Room 1"  # Change for each instance

# Dark theme colors
BG_COLOR = "#121212"        # Very dark background
FG_COLOR = "#00FFFF"        # Bright cyan text
BUTTON_BG = "#1E1E1E"       # Dark button background
BUTTON_FG = "#00FFFF"       # Bright button text
DISABLED_BG = "#333333"
DISABLED_FG = "#555555"
RED_COLOR = "#FF5555"
GREEN_COLOR = "#55FF55"

if not os.path.exists(STATE_FILE):
    with open(STATE_FILE, 'w') as f:
        json.dump({"called_tokens": []}, f)

def pick_preferred_font():
    preferred_fonts = ["Montserrat", "Aptos", "Segoe UI", "Helvetica", "Arial"]
    available = list(tkfont.families())
    for f in preferred_fonts:
        if f in available:
            return f
    return "TkDefaultFont"

class TokenCallerApp:
    def __init__(self, master):
        self.master = master
        self.master.title(f"{COUNTER_NAME} Control Panel")
        self.master.geometry("400x280")
        self.master.configure(bg=BG_COLOR)

        self.font_family = pick_preferred_font()

        heading = tk.Label(master, text=COUNTER_NAME, font=(self.font_family, 18, "bold"),
                           bg=BG_COLOR, fg=FG_COLOR)
        heading.pack(pady=5)

        self.token_data = []
        self.current_token = None
        self.counter_closed = False

        # Display window
        self.display_window = tk.Toplevel(master)
        self.display_window.title(f"{COUNTER_NAME} Display")
        self.display_window.geometry("300x200")
        self.display_window.protocol("WM_DELETE_WINDOW", self.on_display_close)
        self.display_window.configure(bg=BG_COLOR)

        self.display_heading = tk.Label(self.display_window, text="KTech",
                                        font=(self.font_family, 20, "bold"), fg=FG_COLOR, bg=BG_COLOR)
        self.display_heading.pack(pady=(10, 5))

        self.counter_heading = tk.Label(self.display_window, text=COUNTER_NAME,
                                        font=(self.font_family, 14, "bold"), fg=FG_COLOR, bg=BG_COLOR)
        self.counter_heading.pack(pady=(0, 10))

        self.display_label = tk.Label(self.display_window, text="Waiting...",
                                      font=(self.font_family, 24, "bold"), fg=FG_COLOR, bg=BG_COLOR)
        self.display_label.pack(expand=True)

        # Control buttons
        btn_style = {"font": (self.font_family, 14, "bold"),
                     "bg": BUTTON_BG, "fg": BUTTON_FG,
                     "activebackground": "#00AAAA", "activeforeground": "#000000"}

        self.call_button = tk.Button(master, text="Call Next", command=self.call_next, **btn_style)
        self.call_button.pack(pady=5, fill='x')

        btn_style_sm = {"font": (self.font_family, 12, "bold"),
                        "bg": BUTTON_BG, "fg": BUTTON_FG,
                        "activebackground": "#00AAAA", "activeforeground": "#000000"}

        self.recall_button = tk.Button(master, text="Recall", command=self.recall, **btn_style_sm)
        self.recall_button.pack(pady=5, fill='x')

        self.waiting_button = tk.Button(master, text="Waiting", command=self.set_waiting, **btn_style_sm)
        self.waiting_button.pack(pady=5, fill='x')

        self.close_button = tk.Button(master, text="Close Room", fg="white", bg=RED_COLOR,
                                      command=self.close_counter, font=(self.font_family, 12, "bold"))
        self.close_button.pack(pady=5, fill='x')

        self.open_button = tk.Button(master, text="Open Room", fg="white", bg=GREEN_COLOR,
                                     command=self.open_counter, font=(self.font_family, 12, "bold"))
        self.open_button.pack(pady=5, fill='x')

        self.token_label = tk.Label(master, text="Token: -\nName: -", font=(self.font_family, 14, "bold"),
                                    fg=FG_COLOR, bg=BG_COLOR)
        self.token_label.pack(pady=5)

        self.load_tokens()
        self.refresh_excel_data()

    def on_display_close(self):
        messagebox.showinfo("Info", "Display window cannot be closed separately.")

    def load_tokens(self):
        self.token_data = []
        if not os.path.exists(EXCEL_FILE):
            messagebox.showerror("Missing File", f"{EXCEL_FILE} not found.")
            return

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        today = datetime.now().strftime("%Y-%m-%d")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == today:
                self.token_data.append({
                    "token": row[5],  # Corrected from row[7] to row[5]
                    "name": row[3],
                    "date": row[0],
                    "time": row[2]
                })

    def refresh_excel_data(self):
        if not self.counter_closed:
            self.load_tokens()
        self.master.after(3000, self.refresh_excel_data)

    def call_next(self):
        if self.counter_closed:
            messagebox.showwarning("Room Closed", "This room is closed.")
            return

        with open(STATE_FILE, 'r') as f:
            state = json.load(f)

        called_tokens = [item['token'] for item in state.get("called_tokens", [])]

        next_token = next((t for t in self.token_data if t["token"] not in called_tokens), None)

        if next_token:
            self.current_token = next_token
            state["called_tokens"].append({
                "token": next_token["token"],
                "name": next_token["name"],
                "counter": COUNTER_NAME,
                "time": next_token["time"]
            })
            with open(STATE_FILE, 'w') as f:
                json.dump(state, f, indent=2)
            self.update_display(next_token)
        else:
            messagebox.showinfo("Info", "No more tokens to call.")

    def recall(self):
        if self.counter_closed:
            messagebox.showwarning("Room Closed", "This room is closed.")
            return

        if self.current_token:
            self.update_display(self.current_token)
        else:
            messagebox.showinfo("Info", "No token to recall.")

    def set_waiting(self):
        if self.counter_closed:
            messagebox.showwarning("Room Closed", "This room is closed.")
            return

        self.current_token = None
        self.token_label.config(text="Waiting", fg=FG_COLOR)
        self.display_label.config(text="Waiting", fg=FG_COLOR, font=(self.font_family, 24, "bold"))

    def update_display(self, token_info):
        token_text = f"Token: {token_info['token']}\nName: {token_info['name']}"
        self.token_label.config(text=token_text, fg=FG_COLOR)
        self.display_label.config(text=f"Token {token_info['token']}\n{token_info['name']}",
                                  fg=FG_COLOR, font=(self.font_family, 24, "bold"))

    def close_counter(self):
        if messagebox.askyesno("Close Room", "Are you sure you want to close this interview room?"):
            self.counter_closed = True
            self.call_button.config(state='disabled', bg=DISABLED_BG, fg=DISABLED_FG)
            self.recall_button.config(state='disabled', bg=DISABLED_BG, fg=DISABLED_FG)
            self.waiting_button.config(state='disabled', bg=DISABLED_BG, fg=DISABLED_FG)
            self.close_button.config(state='disabled', bg=DISABLED_BG, fg=DISABLED_FG)

            self.display_label.config(text="Room Closed", fg=RED_COLOR, font=(self.font_family, 28, "bold"))
            self.token_label.config(text="Room Closed", fg=RED_COLOR)

    def open_counter(self):
        if not self.counter_closed:
            messagebox.showinfo("Info", "Room is already open.")
            return

        self.counter_closed = False
        self.call_button.config(state='normal', bg=BUTTON_BG, fg=BUTTON_FG)
        self.recall_button.config(state='normal', bg=BUTTON_BG, fg=BUTTON_FG)
        self.waiting_button.config(state='normal', bg=BUTTON_BG, fg=BUTTON_FG)
        self.close_button.config(state='normal', bg=RED_COLOR, fg="white")

        self.current_token = None
        self.token_label.config(text="Waiting", fg=FG_COLOR)
        self.display_label.config(text="Waiting", fg=FG_COLOR, font=(self.font_family, 24, "bold"))


if __name__ == "__main__":
    root = tk.Tk()
    app = TokenCallerApp(root)
    root.mainloop()
